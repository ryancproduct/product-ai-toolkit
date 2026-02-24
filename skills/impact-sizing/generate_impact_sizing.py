#!/usr/bin/env python3
"""
Impact Sizing Excel Generator
Generates a comprehensive impact sizing workbook for translating product hypotheses into ARR estimates.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import pandas as pd
import re
from pathlib import Path


# =============================================================================
# DATA PARSING HELPERS - For processing uploaded files
# =============================================================================

def parse_salesforce_export(file_path: str, feature_keywords: list) -> dict:
    """
    Parse a Salesforce closed sales export and extract feature-related data.

    Args:
        file_path: Path to CSV or Excel file
        feature_keywords: List of keywords to search for in Additional Details

    Returns:
        dict with closed_lost_total, closed_lost_deals, total_deals_value
    """
    # Read file
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)

    # Normalize column names
    df.columns = df.columns.str.strip()

    # Find the right columns (Salesforce exports vary)
    stage_col = next((c for c in df.columns if 'stage' in c.lower()), None)
    reason_col = next((c for c in df.columns if 'reason' in c.lower() or 'churn' in c.lower()), None)
    details_col = next((c for c in df.columns if 'detail' in c.lower() or 'additional' in c.lower()), None)
    amount_col = next((c for c in df.columns if 'price' in c.lower() or 'amount' in c.lower() or 'arr' in c.lower()), None)
    name_col = next((c for c in df.columns if 'account' in c.lower() or 'opportunity' in c.lower() or 'name' in c.lower()), None)

    result = {
        'total_deals_value': 0,
        'closed_lost_total': 0,
        'feature_closed_lost': 0,
        'feature_deals': [],
        'columns_found': {
            'stage': stage_col,
            'reason': reason_col,
            'details': details_col,
            'amount': amount_col,
            'name': name_col
        }
    }

    if not amount_col:
        result['error'] = 'Could not find amount/ARR column'
        return result

    # Convert amount to numeric
    df[amount_col] = pd.to_numeric(df[amount_col].astype(str).str.replace(r'[,$]', '', regex=True), errors='coerce')
    result['total_deals_value'] = df[amount_col].sum()

    # Filter for closed lost
    if stage_col:
        closed_lost = df[df[stage_col].str.contains('Closed Lost', case=False, na=False)]
        result['closed_lost_total'] = closed_lost[amount_col].sum()

        # Further filter by product reason if we have that column
        if reason_col:
            product_lost = closed_lost[closed_lost[reason_col].str.contains('Product', case=False, na=False)]
        else:
            product_lost = closed_lost

        # Search for feature keywords
        if details_col and feature_keywords:
            pattern = '|'.join(feature_keywords)
            feature_related = product_lost[
                product_lost[details_col].str.contains(pattern, case=False, na=False)
            ]
        else:
            feature_related = product_lost

        result['feature_closed_lost'] = feature_related[amount_col].sum()

        # Extract deal details
        for _, row in feature_related.iterrows():
            deal = {
                'name': row.get(name_col, 'Unknown') if name_col else 'Unknown',
                'amount': row.get(amount_col, 0),
                'reason': row.get(reason_col, '') if reason_col else '',
                'details': row.get(details_col, '') if details_col else ''
            }
            result['feature_deals'].append(deal)

    return result


def parse_churn_data(file_path: str, feature_keywords: list) -> dict:
    """
    Parse churn data spreadsheet to find feature-related churned customers.

    Args:
        file_path: Path to CSV or Excel file
        feature_keywords: List of keywords to search for

    Returns:
        dict with churned_customers list and total
    """
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)

    df.columns = df.columns.str.strip()

    # Find columns
    name_col = next((c for c in df.columns if 'account' in c.lower() or 'customer' in c.lower() or 'name' in c.lower()), None)
    arr_col = next((c for c in df.columns if 'arr' in c.lower() or 'revenue' in c.lower() or 'amount' in c.lower()), None)
    reason_col = next((c for c in df.columns if 'reason' in c.lower() or 'churn' in c.lower()), None)

    result = {
        'churned_customers': [],
        'total_churned_arr': 0,
        'feature_churned_arr': 0
    }

    if not arr_col:
        result['error'] = 'Could not find ARR/revenue column'
        return result

    df[arr_col] = pd.to_numeric(df[arr_col].astype(str).str.replace(r'[,$]', '', regex=True), errors='coerce')
    result['total_churned_arr'] = df[arr_col].sum()

    # Search for feature-related churn
    if reason_col and feature_keywords:
        pattern = '|'.join(feature_keywords)
        feature_churned = df[df[reason_col].str.contains(pattern, case=False, na=False)]

        result['feature_churned_arr'] = feature_churned[arr_col].sum()

        for _, row in feature_churned.iterrows():
            customer = {
                'name': row.get(name_col, 'Unknown') if name_col else 'Unknown',
                'arr': row.get(arr_col, 0),
                'reason': row.get(reason_col, '') if reason_col else ''
            }
            result['churned_customers'].append(customer)

    return result


def parse_pasted_table(text: str) -> pd.DataFrame:
    """
    Parse a pasted table (from Confluence, Slack, etc.) into a DataFrame.
    Handles pipe-delimited, tab-delimited, and common table formats.
    """
    lines = text.strip().split('\n')

    # Try to detect delimiter
    if '|' in lines[0]:
        # Pipe-delimited (Markdown/Confluence)
        rows = []
        for line in lines:
            if line.strip() and not re.match(r'^[\s\-|]+$', line):  # Skip separator lines
                cells = [c.strip() for c in line.split('|') if c.strip()]
                if cells:
                    rows.append(cells)
        if rows:
            df = pd.DataFrame(rows[1:], columns=rows[0]) if len(rows) > 1 else pd.DataFrame(rows)
            return df

    elif '\t' in lines[0]:
        # Tab-delimited
        rows = [line.split('\t') for line in lines if line.strip()]
        if rows:
            df = pd.DataFrame(rows[1:], columns=rows[0]) if len(rows) > 1 else pd.DataFrame(rows)
            return df

    # Fallback: comma-separated
    rows = [line.split(',') for line in lines if line.strip()]
    if rows:
        df = pd.DataFrame(rows[1:], columns=rows[0]) if len(rows) > 1 else pd.DataFrame(rows)
        return df

    return pd.DataFrame()


def create_impact_sizing_workbook(data: dict, output_path: str) -> str:
    """
    Create an impact sizing Excel workbook from collected data.

    Args:
        data: Dictionary containing all impact sizing inputs
        output_path: Path to save the Excel file

    Returns:
        Path to the created file
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="4472C4")
    input_font = Font(color="0000FF")  # Blue for inputs
    formula_font = Font(color="000000")  # Black for formulas
    assumption_fill = PatternFill("solid", fgColor="FFFF00")  # Yellow for assumptions
    currency_format = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"_);_(@_)'
    percent_format = '0.0%'
    number_format = '#,##0'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== SUMMARY SHEET ==========
    summary = wb.active
    summary.title = "Summary"

    # Title
    summary['A1'] = data.get('feature_name', 'Impact Sizing')
    summary['A1'].font = Font(bold=True, size=16)
    summary.merge_cells('A1:E1')

    # Hypothesis
    summary['A3'] = "Hypothesis:"
    summary['A3'].font = Font(bold=True)
    summary['B3'] = data.get('hypothesis', '')
    summary.merge_cells('B3:E3')
    summary['B3'].alignment = Alignment(wrap_text=True)
    summary.row_dimensions[3].height = 45

    # Date
    summary['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"

    # Summary table headers
    row = 7
    headers = ['Category', 'Description', 'ARR Impact', 'Confidence', 'Source']
    for col, header in enumerate(headers, 1):
        cell = summary.cell(row=row, column=col, value=header)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Summary data rows
    categories = [
        ('TAM/SAM/SOM (Top-down)', 'topdown_som', 'topdown_source'),
        ('TAM/SAM/SOM (Bottom-up)', 'bottomup_som', 'bottomup_source'),
        ('Churned Customers', 'churned_customers', 'churn_source'),
        ('Tangible Churn Risk', 'churn_risk', 'churn_risk_source'),
        ('Retention Uplift', 'retention_uplift', 'retention_source'),
        ('Closed Won Opportunity', 'closed_won_opportunity', 'opportunity_source'),
        ('Current Pipeline Opportunities', 'current_opportunities', 'pipeline_source'),
    ]

    data_start_row = row + 1
    for i, (cat_name, data_key, source_key) in enumerate(categories):
        r = data_start_row + i
        summary.cell(row=r, column=1, value=cat_name).border = thin_border
        summary.cell(row=r, column=2, value=data.get(f'{data_key}_desc', '')).border = thin_border

        value_cell = summary.cell(row=r, column=3)
        value_cell.value = data.get(data_key, 0)
        value_cell.number_format = currency_format
        value_cell.border = thin_border
        value_cell.font = input_font

        conf_cell = summary.cell(row=r, column=4, value=data.get(f'{data_key}_confidence', 'Medium'))
        conf_cell.border = thin_border

        source_cell = summary.cell(row=r, column=5, value=data.get(source_key, ''))
        source_cell.border = thin_border

    # Total row
    total_row = data_start_row + len(categories)
    summary.cell(row=total_row, column=1, value='TOTAL ARR IMPACT').font = Font(bold=True)
    summary.cell(row=total_row, column=1).border = thin_border
    summary.cell(row=total_row, column=2).border = thin_border

    total_cell = summary.cell(row=total_row, column=3)
    total_cell.value = f'=SUM(C{data_start_row}:C{total_row-1})'
    total_cell.number_format = currency_format
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border

    summary.cell(row=total_row, column=4).border = thin_border
    summary.cell(row=total_row, column=5).border = thin_border

    # Column widths
    summary.column_dimensions['A'].width = 30
    summary.column_dimensions['B'].width = 40
    summary.column_dimensions['C'].width = 18
    summary.column_dimensions['D'].width = 12
    summary.column_dimensions['E'].width = 50

    # ========== TAM/SAM/SOM SHEET ==========
    tam_sheet = wb.create_sheet("TAM_SAM_SOM")

    tam_sheet['A1'] = "TAM/SAM/SOM Analysis"
    tam_sheet['A1'].font = header_font

    # Top-down section
    tam_sheet['A3'] = "TOP-DOWN APPROACH (Industry Data)"
    tam_sheet['A3'].font = section_font
    tam_sheet['A3'].fill = section_fill
    tam_sheet.merge_cells('A3:E3')

    topdown_rows = [
        ('TAM (Total Addressable Market)', 'tam_topdown', 'Total market opportunity globally'),
        ('SAM (Serviceable Addressable Market)', 'sam_topdown', 'Market within strategic focus'),
        ('Target Customers in SAM', 'customers_topdown', 'Number of potential customers'),
        ('Capture Rate Assumption', 'capture_rate_topdown', 'Realistic % we can win'),
        ('Average Seats per Customer', 'avg_seats', 'Typical deployment size'),
        ('Price per Seat (Annual)', 'seat_price', 'Annual seat cost'),
        ('Win Rate', 'win_rate', 'Current closed-won rate'),
        ('SOM (Serviceable Obtainable Market)', 'topdown_som', 'Realistic ARR opportunity'),
    ]

    row = 5
    for label, key, desc in topdown_rows:
        tam_sheet.cell(row=row, column=1, value=label)

        val_cell = tam_sheet.cell(row=row, column=2)
        val_cell.value = data.get(key, 0)
        if 'rate' in key.lower():
            val_cell.number_format = percent_format
        elif key in ['customers_topdown', 'avg_seats']:
            val_cell.number_format = number_format
        else:
            val_cell.number_format = currency_format
        val_cell.font = input_font

        tam_sheet.cell(row=row, column=3, value=desc)
        tam_sheet.cell(row=row, column=4, value=data.get(f'{key}_source', ''))
        row += 1

    # SOM Formula row
    som_row = row
    tam_sheet.cell(row=som_row, column=1, value="SOM Calculation:")
    tam_sheet.cell(row=som_row, column=2, value="= Customers × Capture Rate × Seats × Price × Win Rate")

    # Bottom-up section
    row += 3
    tam_sheet.cell(row=row, column=1, value="BOTTOM-UP APPROACH (Internal Data)")
    tam_sheet[f'A{row}'].font = section_font
    tam_sheet[f'A{row}'].fill = section_fill
    tam_sheet.merge_cells(f'A{row}:E{row}')

    row += 2
    bottomup_rows = [
        ('TAM (Total orgs in focus industries)', 'tam_bottomup', 'From Industry Playbooks'),
        ('Feature Requests (this feature)', 'feature_requests', 'From Productboard'),
        ('Total Feature Requests (product area)', 'total_requests', 'Total requests in product area'),
        ('SAM % (feature/total requests)', 'sam_percent', 'Demand proportion'),
        ('SAM (Orgs)', 'sam_bottomup', 'TAM × SAM %'),
        ('Capture Rate', 'capture_rate_bottomup', 'Realistic capture'),
        ('Average Seats', 'avg_seats_bottomup', 'Seats per customer'),
        ('Price per Seat (Annual)', 'seat_price_bottomup', 'Annual seat cost'),
        ('Win Rate', 'win_rate_bottomup', 'Closed-won rate'),
        ('Deal Cycle Adjustment', 'deal_cycle_adj', 'Based on 89-day cycle'),
        ('SOM (Bottom-up)', 'bottomup_som', 'Realistic ARR opportunity'),
    ]

    for label, key, desc in bottomup_rows:
        tam_sheet.cell(row=row, column=1, value=label)

        val_cell = tam_sheet.cell(row=row, column=2)
        val_cell.value = data.get(key, 0)
        if 'rate' in key.lower() or 'percent' in key.lower() or 'adj' in key.lower():
            val_cell.number_format = percent_format
        elif key in ['feature_requests', 'total_requests', 'sam_bottomup', 'avg_seats_bottomup']:
            val_cell.number_format = number_format
        else:
            val_cell.number_format = currency_format
        val_cell.font = input_font

        tam_sheet.cell(row=row, column=3, value=desc)
        row += 1

    tam_sheet.column_dimensions['A'].width = 35
    tam_sheet.column_dimensions['B'].width = 20
    tam_sheet.column_dimensions['C'].width = 35
    tam_sheet.column_dimensions['D'].width = 50

    # ========== CHURN SHEET ==========
    churn_sheet = wb.create_sheet("Churn_Analysis")

    churn_sheet['A1'] = "Churn Analysis"
    churn_sheet['A1'].font = header_font

    # Churned customers section
    churn_sheet['A3'] = "CHURNED CUSTOMERS"
    churn_sheet['A3'].font = section_font
    churn_sheet['A3'].fill = section_fill
    churn_sheet.merge_cells('A3:E3')

    churn_sheet['A5'] = "Customer Name"
    churn_sheet['B5'] = "ARR"
    churn_sheet['C5'] = "Churn Reason"
    churn_sheet['D5'] = "Source"
    for col in range(1, 5):
        churn_sheet.cell(row=5, column=col).font = Font(bold=True)

    # Add churned customer data
    churned_customers = data.get('churned_customers_list', [])
    row = 6
    for customer in churned_customers:
        churn_sheet.cell(row=row, column=1, value=customer.get('name', ''))
        arr_cell = churn_sheet.cell(row=row, column=2, value=customer.get('arr', 0))
        arr_cell.number_format = currency_format
        arr_cell.font = input_font
        churn_sheet.cell(row=row, column=3, value=customer.get('reason', ''))
        churn_sheet.cell(row=row, column=4, value=customer.get('source', ''))
        row += 1

    churn_total_row = row
    churn_sheet.cell(row=churn_total_row, column=1, value="TOTAL CHURNED").font = Font(bold=True)
    churn_sheet.cell(row=churn_total_row, column=2, value=f'=SUM(B6:B{row-1})')
    churn_sheet.cell(row=churn_total_row, column=2).number_format = currency_format
    churn_sheet.cell(row=churn_total_row, column=2).font = Font(bold=True)

    # Tangible churn risk section
    row += 3
    churn_sheet.cell(row=row, column=1, value="TANGIBLE CHURN RISK")
    churn_sheet[f'A{row}'].font = section_font
    churn_sheet[f'A{row}'].fill = section_fill
    churn_sheet.merge_cells(f'A{row}:E{row}')

    row += 2
    churn_sheet.cell(row=row, column=1, value="Customer Name").font = Font(bold=True)
    churn_sheet.cell(row=row, column=2, value="ARR at Risk").font = Font(bold=True)
    churn_sheet.cell(row=row, column=3, value="Risk Details").font = Font(bold=True)
    churn_sheet.cell(row=row, column=4, value="Source (Twine/Slack)").font = Font(bold=True)

    risk_start = row + 1
    churn_risks = data.get('churn_risk_list', [])
    row += 1
    for risk in churn_risks:
        churn_sheet.cell(row=row, column=1, value=risk.get('name', ''))
        arr_cell = churn_sheet.cell(row=row, column=2, value=risk.get('arr', 0))
        arr_cell.number_format = currency_format
        arr_cell.font = input_font
        churn_sheet.cell(row=row, column=3, value=risk.get('details', ''))
        churn_sheet.cell(row=row, column=4, value=risk.get('source', ''))
        row += 1

    churn_sheet.cell(row=row, column=1, value="TOTAL AT RISK").font = Font(bold=True)
    churn_sheet.cell(row=row, column=2, value=f'=SUM(B{risk_start}:B{row-1})')
    churn_sheet.cell(row=row, column=2).number_format = currency_format
    churn_sheet.cell(row=row, column=2).font = Font(bold=True)

    # Retention uplift section
    row += 3
    churn_sheet.cell(row=row, column=1, value="RETENTION UPLIFT CALCULATION")
    churn_sheet[f'A{row}'].font = section_font
    churn_sheet[f'A{row}'].fill = section_fill
    churn_sheet.merge_cells(f'A{row}:E{row}')

    row += 2
    retention_inputs = [
        ('FY25 Total Company Churn', data.get('fy25_total_churn', 19600000), currency_format),
        ('Churned Due to This Feature', data.get('churned_customers', 0), currency_format),
        ('% of Total Churn', '=B{}/B{}'.format(row+1, row), percent_format),
        ('FY26 GDR Uplift Target (%)', data.get('gdr_target', 0.02), percent_format),
        ('FY26 Churn Baseline', '=B{}*(1-B{})'.format(row, row+3), currency_format),
        ('Expected Retention Uplift', '=B{}*B{}'.format(row+2, row+4), currency_format),
    ]

    for label, value, fmt in retention_inputs:
        churn_sheet.cell(row=row, column=1, value=label)
        val_cell = churn_sheet.cell(row=row, column=2, value=value)
        val_cell.number_format = fmt
        if not str(value).startswith('='):
            val_cell.font = input_font
        row += 1

    churn_sheet.column_dimensions['A'].width = 30
    churn_sheet.column_dimensions['B'].width = 18
    churn_sheet.column_dimensions['C'].width = 40
    churn_sheet.column_dimensions['D'].width = 50

    # ========== OPPORTUNITIES SHEET ==========
    opp_sheet = wb.create_sheet("Opportunities")

    opp_sheet['A1'] = "Opportunities Analysis"
    opp_sheet['A1'].font = header_font

    # Closed Lost Analysis
    opp_sheet['A3'] = "CLOSED LOST ANALYSIS"
    opp_sheet['A3'].font = section_font
    opp_sheet['A3'].fill = section_fill
    opp_sheet.merge_cells('A3:E3')

    row = 5
    closed_lost_inputs = [
        ('Total Closed Deals (2 FY)', data.get('total_closed_deals', 141000000), currency_format),
        ('Feature-Related Closed Lost', data.get('feature_closed_lost', 0), currency_format),
        ('% of Total', '', percent_format),  # Formula
        ('FY26 New Bookings Target', data.get('fy26_bookings_target', 63000000), currency_format),
        ('Expected Bookings Uplift', '', currency_format),  # Formula
    ]

    for i, (label, value, fmt) in enumerate(closed_lost_inputs):
        opp_sheet.cell(row=row, column=1, value=label)
        val_cell = opp_sheet.cell(row=row, column=2)
        if i == 2:  # % of Total formula
            val_cell.value = f'=B{row-1}/B{row-2}'
        elif i == 4:  # Expected uplift formula
            val_cell.value = f'=B{row-2}*B{row-1}'
        else:
            val_cell.value = value
            val_cell.font = input_font
        val_cell.number_format = fmt
        row += 1

    # Current Pipeline
    row += 2
    opp_sheet.cell(row=row, column=1, value="CURRENT PIPELINE OPPORTUNITIES")
    opp_sheet[f'A{row}'].font = section_font
    opp_sheet[f'A{row}'].fill = section_fill
    opp_sheet.merge_cells(f'A{row}:E{row}')

    row += 2
    opp_sheet.cell(row=row, column=1, value="Opportunity Name").font = Font(bold=True)
    opp_sheet.cell(row=row, column=2, value="Value").font = Font(bold=True)
    opp_sheet.cell(row=row, column=3, value="Stage").font = Font(bold=True)
    opp_sheet.cell(row=row, column=4, value="Notes").font = Font(bold=True)

    pipeline_start = row + 1
    pipeline_opps = data.get('pipeline_opportunities', [])
    row += 1
    for opp in pipeline_opps:
        opp_sheet.cell(row=row, column=1, value=opp.get('name', ''))
        val_cell = opp_sheet.cell(row=row, column=2, value=opp.get('value', 0))
        val_cell.number_format = currency_format
        val_cell.font = input_font
        opp_sheet.cell(row=row, column=3, value=opp.get('stage', ''))
        opp_sheet.cell(row=row, column=4, value=opp.get('notes', ''))
        row += 1

    opp_sheet.cell(row=row, column=1, value="TOTAL PIPELINE").font = Font(bold=True)
    opp_sheet.cell(row=row, column=2, value=f'=SUM(B{pipeline_start}:B{row-1})')
    opp_sheet.cell(row=row, column=2).number_format = currency_format
    opp_sheet.cell(row=row, column=2).font = Font(bold=True)

    opp_sheet.column_dimensions['A'].width = 35
    opp_sheet.column_dimensions['B'].width = 18
    opp_sheet.column_dimensions['C'].width = 20
    opp_sheet.column_dimensions['D'].width = 40

    # ========== DATA SOURCES SHEET ==========
    sources_sheet = wb.create_sheet("Data_Sources")

    sources_sheet['A1'] = "Data Sources & References"
    sources_sheet['A1'].font = header_font

    sources_sheet['A3'] = "Description"
    sources_sheet['B3'] = "Link/Reference"
    sources_sheet['C3'] = "Access Date"
    sources_sheet['D3'] = "Notes"
    for col in range(1, 5):
        sources_sheet.cell(row=3, column=col).font = Font(bold=True)
        sources_sheet.cell(row=3, column=col).fill = section_fill
        sources_sheet.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")

    # Standard data sources — customise these URLs for your company
    standard_sources = [
        ('Average seat sizes by segment', '[Your internal resource URL]', '', 'Add your URL'),
        ('Closed Sales Report', '[Your CRM report URL]', '', 'E.g. Salesforce, HubSpot'),
        ('All Customers Report', '[Your BI dashboard URL]', '', 'E.g. Tableau, Looker'),
        ('Product Churn Deep-dive', '[Your churn analysis URL]', '', 'Add your URL'),
        ('Industry/Segment Playbooks', '[Your internal docs]', '', ''),
        ('Feature Requests', '[Your feedback tool URL]', '', 'E.g. Productboard, Canny'),
        ('Customer Insights', '[Your customer intel tool URL]', '', 'E.g. Twine, Gong'),
        ('Account Risks Channel', '[Your Slack/Teams channel]', '', ''),
    ]

    row = 4
    for desc, link, date, notes in standard_sources:
        sources_sheet.cell(row=row, column=1, value=desc)
        sources_sheet.cell(row=row, column=2, value=link)
        sources_sheet.cell(row=row, column=3, value=date)
        sources_sheet.cell(row=row, column=4, value=notes)
        row += 1

    # Add custom sources from data
    custom_sources = data.get('custom_sources', [])
    for source in custom_sources:
        sources_sheet.cell(row=row, column=1, value=source.get('description', ''))
        sources_sheet.cell(row=row, column=2, value=source.get('link', ''))
        sources_sheet.cell(row=row, column=3, value=source.get('date', ''))
        sources_sheet.cell(row=row, column=4, value=source.get('notes', ''))
        row += 1

    sources_sheet.column_dimensions['A'].width = 35
    sources_sheet.column_dimensions['B'].width = 60
    sources_sheet.column_dimensions['C'].width = 15
    sources_sheet.column_dimensions['D'].width = 30

    # ========== ASSUMPTIONS SHEET ==========
    assumptions_sheet = wb.create_sheet("Assumptions")

    assumptions_sheet['A1'] = "Key Assumptions"
    assumptions_sheet['A1'].font = header_font

    assumptions_sheet['A3'] = "Assumption"
    assumptions_sheet['B3'] = "Value"
    assumptions_sheet['C3'] = "Rationale"
    assumptions_sheet['D3'] = "Confidence"
    for col in range(1, 5):
        assumptions_sheet.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
        assumptions_sheet.cell(row=3, column=col).fill = section_fill

    # Standard assumptions
    assumptions = [
        ('Capture Rate (new customers)', '5%', 'Conservative estimate for new market entry', 'Medium'),
        ('Average Seats per Customer', '5', 'Based on current customer base average', 'High'),
        ('Standard Seat Price (Annual)', '$288', '$24/month × 12', 'High'),
        ('Enterprise Seat Price (Annual)', '$367', 'Blended rate from All Customers Report', 'High'),
        ('Win Rate', '23%', 'From Global Closed Sales Report', 'High'),
        ('Average Deal Cycle', '89 days', 'From Global Closed Sales Report', 'High'),
        ('New Bookings Target', '[Your target]', 'Company Strategy', 'High'),
        ('Prior Year Total Churn', '[Your figure]', 'Churn Analysis', 'High'),
        ('GDR Uplift Target', '[Your target]', 'Company Strategy', 'High'),
    ]

    row = 4
    for assumption, value, rationale, confidence in assumptions:
        assumptions_sheet.cell(row=row, column=1, value=assumption)
        assumptions_sheet.cell(row=row, column=2, value=value)
        assumptions_sheet.cell(row=row, column=2).fill = assumption_fill
        assumptions_sheet.cell(row=row, column=3, value=rationale)
        assumptions_sheet.cell(row=row, column=4, value=confidence)
        row += 1

    # Custom assumptions
    custom_assumptions = data.get('custom_assumptions', [])
    for assumption in custom_assumptions:
        assumptions_sheet.cell(row=row, column=1, value=assumption.get('name', ''))
        assumptions_sheet.cell(row=row, column=2, value=assumption.get('value', ''))
        assumptions_sheet.cell(row=row, column=2).fill = assumption_fill
        assumptions_sheet.cell(row=row, column=3, value=assumption.get('rationale', ''))
        assumptions_sheet.cell(row=row, column=4, value=assumption.get('confidence', 'Medium'))
        row += 1

    assumptions_sheet.column_dimensions['A'].width = 35
    assumptions_sheet.column_dimensions['B'].width = 15
    assumptions_sheet.column_dimensions['C'].width = 50
    assumptions_sheet.column_dimensions['D'].width = 12

    # Save workbook
    wb.save(output_path)
    return output_path


def create_blank_template(output_path: str) -> str:
    """Create a blank impact sizing template with placeholder values."""

    template_data = {
        'feature_name': '[Feature Name] - Impact Sizing',
        'hypothesis': 'By <making a change> we will see <metric> <increase/decrease> because <customer problem solved>',

        # TAM/SAM/SOM Top-down
        'tam_topdown': 0,
        'sam_topdown': 0,
        'customers_topdown': 0,
        'capture_rate_topdown': 0.05,
        'avg_seats': 5,
        'seat_price': 288,
        'win_rate': 0.23,
        'topdown_som': 0,

        # TAM/SAM/SOM Bottom-up
        'tam_bottomup': 0,
        'feature_requests': 0,
        'total_requests': 0,
        'sam_percent': 0,
        'sam_bottomup': 0,
        'capture_rate_bottomup': 0.05,
        'avg_seats_bottomup': 5,
        'seat_price_bottomup': 367,
        'win_rate_bottomup': 0.23,
        'deal_cycle_adj': 1.0,
        'bottomup_som': 0,

        # Churn
        'churned_customers': 0,
        'churned_customers_list': [],
        'churn_risk': 0,
        'churn_risk_list': [],
        'fy25_total_churn': 19600000,
        'gdr_target': 0.02,
        'retention_uplift': 0,

        # Opportunities
        'total_closed_deals': 141000000,
        'feature_closed_lost': 0,
        'fy26_bookings_target': 63000000,
        'closed_won_opportunity': 0,
        'pipeline_opportunities': [],
        'current_opportunities': 0,

        # Sources
        'custom_sources': [],
        'custom_assumptions': [],
    }

    return create_impact_sizing_workbook(template_data, output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python generate_impact_sizing.py <output_path> [data_json]")
        print("\nCreates an impact sizing Excel workbook.")
        print("\nIf data_json is not provided, creates a blank template.")
        sys.exit(1)

    output_path = sys.argv[1]

    if len(sys.argv) > 2:
        data = json.loads(sys.argv[2])
        create_impact_sizing_workbook(data, output_path)
    else:
        create_blank_template(output_path)

    print(f"Created: {output_path}")
